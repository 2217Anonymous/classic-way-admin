"""Mark integrated VS stories as Done and regenerate tracker."""

from __future__ import annotations

import re
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent
STORIES = ROOT / "tracker_stories.py"

DONE_VS = [
    f"VS-{i:03d}" for i in range(1, 29)
]


def main() -> None:
    text = STORIES.read_text(encoding="utf-8")
    for sid in DONE_VS:
        text = re.sub(
            rf'(\("{re.escape(sid)}",[\s\S]*?"(?:P0|P1|P2)",\s*")(?:Planned|In Progress)(")',
            r"\1Done\2",
            text,
            count=1,
        )
    for wf in ["WF-A", "WF-B", "WF-C", "WF-D", "WF-E", "WF-F"]:
        text = re.sub(
            rf'(\("{re.escape(wf)}",[\s\S]*?"(?:P0|P1)",\s*")(?:Planned|In Progress)("\))',
            r"\1Done\2",
            text,
            count=1,
        )
    # S* timeline phases
    for phase in [
        "S1 Theme",
        "S2 Catalog Sync",
        "S3 Commerce",
        "S4 Admin Reflect",
        "S5 Fulfillment",
        "S6 Harden",
    ]:
        text = re.sub(
            rf'(\("{re.escape(phase)}",[\s\S]*?)(?:Planned|In Progress)("\))',
            r"\1Done\2",
            text,
            count=1,
        )
    STORIES.write_text(text, encoding="utf-8")

    from generate_client_materials import build_excel

    xlsx = ROOT / "PROJECT_TRACKING.xlsx"
    build_excel(xlsx)
    wb = load_workbook(xlsx)
    rows = list(wb["Stories"].iter_rows(min_row=2, values_only=True))
    print("stories", len(rows), dict(Counter(r[10] for r in rows)))
    print(
        "VS Done",
        sum(1 for r in rows if str(r[0]).startswith("VS-") and r[10] == "Done"),
    )


if __name__ == "__main__":
    main()
